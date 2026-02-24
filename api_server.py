"""
Serwer API dla Systemu Obiegu Kopert.
Łączy frontend (prototype.html) z bazą SQLite.

Uruchomienie: python3 api_server.py
Serwer dostępny na: http://localhost:5000
"""
from flask import Flask, jsonify, request, send_file, Response
from flask_cors import CORS
from database import db
from domain import EnvelopeStatus, HolderType, CreationReason, Envelope
import json
from datetime import datetime
import os
import hashlib
import hmac
import io
import time
import uuid
from collections import defaultdict, deque
from pathlib import Path
from PIL import Image, UnidentifiedImageError

# --- Kody błędów zgodne ze specyfikacją v2.0 ---
ERROR_CODES = {
    'ERR_DUPLICATE_ACTIVE': 'Koperta aktywna na produkcji!',
    'ERR_NOT_ISSUED': 'Koperta nie została wydana z magazynu!',
    'ERR_WRONG_COLOR': 'Koperta niekompletna (czerwona)!',
    'ERR_MACHINE_BUSY': 'Koperta zajęta przez inną maszynę!',
    'ERR_INVALID_STATUS': 'Nieprawidłowy status koperty!',
    'ERR_NOT_FOUND': 'Koperta nie istnieje w systemie!'
}

app = Flask(__name__)
CORS(app)  # Pozwala na zapytania z przeglądarki (prototype.html)

APP_ENV = os.environ.get('APP_ENV', 'development').lower()
IMAGE_SIGNING_SECRET = os.environ.get('IMAGE_SIGNING_SECRET', '')
if APP_ENV == 'production' and not IMAGE_SIGNING_SECRET:
    raise RuntimeError("Missing IMAGE_SIGNING_SECRET in production")
if not IMAGE_SIGNING_SECRET:
    IMAGE_SIGNING_SECRET = "dev-only-change-me"

NOTE_IMAGES_DIR = Path(os.environ.get('NOTE_IMAGES_DIR', './data/note_images'))
NOTE_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
SIGNED_URL_TTL_SECONDS = int(os.environ.get('SIGNED_URL_TTL_SECONDS', '600'))
MAX_IMAGE_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_IMAGES_PER_NOTE = 3
app.config['MAX_CONTENT_LENGTH'] = MAX_IMAGE_UPLOAD_BYTES + 1024 * 1024

_rate_limit_store: dict[str, deque[float]] = defaultdict(deque)
_note_image_metrics: dict[str, int] = defaultdict(int)

# ========================
# POMOCNICZE FUNKCJE
# ========================

# log_error przeniesione do database.py
# (Zachowujemy puste lub usuwamy, tutaj usuwam definicję bo jest nieużywana w nowym kodzie)

def get_all_envelopes_from_db():
    """Pobiera wszystkie koperty z bazy SQLite wraz z danymi produktu."""
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT e.unique_key, e.rcs_id, e.status, 
               e.current_holder_id, e.current_holder_type, e.warehouse_section,
               e.is_green, e.last_operator_id, e.product_id,
               p.company_name, p.product_name
        FROM envelopes e
        LEFT JOIN products p ON e.product_id = p.id
    """)
    rows = cursor.fetchall()
    conn.close()
    
    result = []
    for row in rows:
        # Jeśli koperta ma przypisany produkt, używamy danych z tabeli products
        if row['product_id'] and row['company_name']:
            product_display = f"{row['company_name']} | {row['product_name']}"
        else:
            product_display = f"KOPERTA {row['rcs_id']}"
        
        result.append({
            "id": row['unique_key'],
            "rcs_id": row['rcs_id'],
            "product": product_display,
            "company_name": row['company_name'],
            "product_name": row['product_name'],
            "status": row['status'],
            "machine": row['current_holder_id'] if row['current_holder_id'] else None,
            "location": row['warehouse_section'] if row['status'] == 'MAGAZYN' else None
        })
    return result

def init_demo_envelopes():
    """Inicjalizuje bazę przykładowymi kopertami jeśli jest pusta."""
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM envelopes")
    count = cursor.fetchone()[0]
    
    if count == 0:
        print("📦 Inicjalizacja bazy demo kopertami...")
        demo_envelopes = [
            ("111222333#1.0#01", "111222333", "1.0", 1, "MAGAZYN", "MAGAZYN_A", "WAREHOUSE", "Sekcja A"),
            ("111222333#1.0#02", "111222333", "1.0", 2, "MAGAZYN", "MAGAZYN_A", "WAREHOUSE", "Sekcja A"),
            ("222333444#2.5#15", "222333444", "2.5", 15, "MAGAZYN", "MAGAZYN_B", "WAREHOUSE", "Sekcja B"),
            ("333444555#1.3#08", "333444555", "1.3", 8, "W_PRODUKCJI", "BOOBST 1", "MACHINE", None),
            ("444555666#3.0#22", "444555666", "3.0", 22, "MAGAZYN", "MAGAZYN_D", "WAREHOUSE", "Sekcja D"),
            ("555666777#1.1#03", "555666777", "1.1", 3, "MAGAZYN", "MAGAZYN_E", "WAREHOUSE", "Sekcja E"),
            ("765432345#1.2#45", "765432345", "1.2", 45, "MAGAZYN", "MAGAZYN_A", "WAREHOUSE", "Sekcja A"),
            ("765432345#1.2#46", "765432345", "1.2", 46, "W_PRODUKCJI", "PRINTER MAIN", "MACHINE", None),
        ]
        
        for env in demo_envelopes:
            cursor.execute("""
                INSERT INTO envelopes 
                (unique_key, rcs_id, product_version, additional_number, status, current_holder_id, current_holder_type, warehouse_section, creation_reason)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'NEW')
            """, env)
        
        conn.commit()
        print(f"✅ Dodano {len(demo_envelopes)} kopert demo.")
    
    conn.close()


def _is_rate_limited(key: str, max_requests: int, window_seconds: int) -> bool:
    now = time.time()
    bucket = _rate_limit_store[key]
    while bucket and now - bucket[0] > window_seconds:
        bucket.popleft()
    if len(bucket) >= max_requests:
        return True
    bucket.append(now)
    return False


def _increment_note_metric(metric_name: str) -> None:
    _note_image_metrics[metric_name] += 1


def _supported_image_magic(data: bytes) -> bool:
    if data.startswith(b'\xFF\xD8\xFF'):  # JPEG
        return True
    if data.startswith(b'\x89PNG\r\n\x1a\n'):  # PNG
        return True
    if data.startswith(b'GIF87a') or data.startswith(b'GIF89a'):  # GIF
        return True
    if len(data) > 12 and data[0:4] == b'RIFF' and data[8:12] == b'WEBP':  # WEBP
        return True
    return False


def _normalize_image_to_webp(file_storage):
    raw = file_storage.read(MAX_IMAGE_UPLOAD_BYTES + 1)
    if not raw:
        return None, "Pusty plik", 400
    if len(raw) > MAX_IMAGE_UPLOAD_BYTES:
        return None, "Plik przekracza limit 10MB", 413
    if not _supported_image_magic(raw):
        return None, "Nieprawidlowy format obrazu (magic bytes)", 400

    try:
        image = Image.open(io.BytesIO(raw))
        image.load()
    except UnidentifiedImageError:
        return None, "Nie udalo sie odczytac obrazu", 400
    except Exception as e:
        return None, f"Uszkodzony obraz: {e}", 400

    if image.mode not in ("RGB", "RGBA"):
        image = image.convert("RGB")

    max_dim = 1600
    image.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)
    if image.mode == "RGBA":
        bg = Image.new("RGB", image.size, (255, 255, 255))
        bg.paste(image, mask=image.split()[3])
        image = bg

    out = io.BytesIO()
    image.save(out, format='WEBP', quality=82, method=6)
    webp_bytes = out.getvalue()
    sha256_hex = hashlib.sha256(webp_bytes).hexdigest()

    return {
        "bytes": webp_bytes,
        "mime_type": "image/webp",
        "width": image.width,
        "height": image.height,
        "size_bytes": len(webp_bytes),
        "sha256": sha256_hex,
    }, None, 200


def _sign_image_token(image_id: int, exp: int) -> str:
    payload = f"{image_id}:{exp}".encode("utf-8")
    return hmac.new(IMAGE_SIGNING_SECRET.encode("utf-8"), payload, hashlib.sha256).hexdigest()


def _build_signed_image_url(image_id: int, exp: int) -> str:
    sig = _sign_image_token(image_id, exp)
    return f"/api/note-images/{image_id}/file?exp={exp}&sig={sig}"


def _verify_signed_image_url(image_id: int, exp_raw: str, sig: str) -> bool:
    try:
        exp = int(exp_raw)
    except Exception:
        return False

    if exp < int(time.time()):
        return False

    expected = _sign_image_token(image_id, exp)
    return hmac.compare_digest(expected, sig or "")


def _serialize_image_meta(image_row: dict) -> dict:
    exp = int(time.time()) + SIGNED_URL_TTL_SECONDS
    return {
        "id": image_row["id"],
        "original_filename": image_row.get("original_filename"),
        "mime_type": image_row["mime_type"],
        "width": image_row["width"],
        "height": image_row["height"],
        "size_bytes": image_row["size_bytes"],
        "order_index": image_row["order_index"],
        "revision": image_row["revision"],
        "annotations_json": image_row.get("annotations_json") or {"objects": []},
        "etag": image_row.get("sha256"),
        "signed_url": _build_signed_image_url(image_row["id"], exp),
        "expires_at": exp,
    }


def _parse_note_cursor(cursor_raw: str):
    if not cursor_raw:
        return None
    parts = cursor_raw.split(",", 1)
    if len(parts) != 2:
        return None
    return cursor_raw

# ========================
# ENDPOINTY API
# ========================

@app.route('/')
def serve_frontend():
    return send_file('prototype.html')

@app.route('/<path:filename>')
def serve_static(filename):
    """Serve static files (.js, .css, .html etc.)"""
    if filename.endswith(('.js', '.css', '.html')):
        return send_file(filename)
    return jsonify({"error": "File not found"}), 404


@app.route('/api/envelopes', methods=['GET'])
def get_all_envelopes():
    """Zwraca listę kopert z paginacją."""
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 50, type=int)
    
    result = db.get_envelopes_paginated(page, limit)
    return jsonify(result)

@app.route('/api/envelopes/search', methods=['GET'])
def search_envelopes():
    """Wyszukuje koperty po fragmencie ID, z opcjonalnym filtrowaniem statusu."""
    query = request.args.get('q', '')
    status_filter = request.args.get('status')
    
    if len(query) < 2:
        return jsonify([])
    
    conn = db.get_connection()
    cursor = conn.cursor()
    
    sql = """
        SELECT unique_key, rcs_id, product_version, status, 
               current_holder_id, current_holder_type, warehouse_section
        FROM envelopes
        WHERE unique_key LIKE ?
    """
    params = [f"%{query}%"]
    
    if status_filter:
        sql += " AND status = ?"
        params.append(status_filter)
        
    sql += " LIMIT 20"
    
    cursor.execute(sql, tuple(params))
    rows = cursor.fetchall()
    conn.close()
    
    results = []
    for row in rows:
        status = row['status']
        holder = row['current_holder_id']
        
        results.append({
            "id": row['unique_key'],
            "product": f"KOPERTA {row['rcs_id']} v{row['product_version']}",
            "status": status,
            "machine": holder if holder else None,
            "location": row['warehouse_section'] if status == 'MAGAZYN' else None
        })
    
    return jsonify(results)

@app.route('/api/envelopes/<path:envelope_id>/status', methods=['GET'])
def get_envelope_status(envelope_id):
    """Pobiera aktualny status konkretnej koperty."""
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT unique_key, status, current_holder_id, current_holder_type, warehouse_section
        FROM envelopes WHERE unique_key = ?
    """, (envelope_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Koperta nie znaleziona"}), 404
    
    return jsonify({
        "id": row['unique_key'],
        "status": row['status'],
        "machine": row['current_holder_id'] if row['current_holder_id'] else None,
        "location": row['warehouse_section']
    })

@app.route('/api/envelopes/<path:envelope_id>/issue', methods=['POST'])
def issue_envelope_from_warehouse(envelope_id):
    """
    Wydaje kopertę z magazynu na wózek transportowy.
    Zmiana statusu: MAGAZYN -> SHOP_FLOOR (oznacza wydanie na produkcję)
    
    Body JSON: { "cart_id": "CART-OUT-1", "user_id": "magazynier1" }
    
    Implementuje blokadę duplikatów wg specyfikacji v2.0:
    - Walidacja statusu (tylko MAGAZYN dozwolony)
    - Walidacja is_green (tylko kompletne koperty)
    - Logowanie błędów do error_logs
    """
    data = request.get_json() or {}
    cart_id = data.get('cart_id', 'CART-OUT')
    user_id = data.get('user_id', 'UNKNOWN')
    
    result = db.issue_envelope(envelope_id, cart_id, user_id)
    
    if result.get("success"):
        return jsonify({
            "success": True,
            "message": f"Koperta {envelope_id} wydana na wózek {cart_id} (SHOP_FLOOR)",
            "status": "SHOP_FLOOR",
            "cart_id": cart_id,
            "operator": user_id
        })
    else:
        status_code = result.pop('status', 400)
        # Mapowanie pola error_code na error message
        if 'error' not in result and 'error_code' in result:
             result['error'] = ERROR_CODES.get(result['error_code'], 'Błąd wydania')
        return jsonify(result), status_code

@app.route('/api/envelopes/<path:envelope_id>/load', methods=['POST'])
def load_envelope_to_machine(envelope_id):
    """
    Ładuje lub automatycznie transferuje kopertę na maszynę.
    Body JSON: { "machine": "BOOBST 1", "operator_id": "BOOBST 1" }
    """
    data = request.get_json() or {}
    machine = data.get('machine', 'UNKNOWN')
    operator_id = data.get('operator_id') or machine
    
    result = db.bind_envelope_to_machine(envelope_id, machine, operator_id)
    
    if result.get("success"):
        op = result.get("operation")
        if op == "TRANSFER_AUTO":
            message = f"Koperta {envelope_id} automatycznie przeniesiona: {result.get('from_machine')} -> {machine}"
        elif op == "ALREADY_ON_MACHINE":
            message = f"Koperta {envelope_id} jest już na maszynie {machine}"
        else:
            message = f"Koperta {envelope_id} załadowana na {machine}"

        return jsonify({
            "success": True,
            "message": message,
            "status": "W_PRODUKCJI",
            "machine": machine,
            "operator": operator_id,
            "operation": op,
            "from_machine": result.get("from_machine")
        })
    else:
        status_code = result.pop('status', 400)
        if 'error' not in result and 'error_code' in result:
             result['error'] = ERROR_CODES.get(result['error_code'], 'Błąd ładowania')
        return jsonify(result), status_code

@app.route('/api/envelopes/<path:envelope_id>/transfer-auto', methods=['POST'])
def transfer_envelope_auto(envelope_id):
    """Alias dla automatycznego transferu/załadowania koperty na maszynę."""
    return load_envelope_to_machine(envelope_id)

@app.route('/api/envelopes/<path:envelope_id>', methods=['DELETE'])
def delete_envelope(envelope_id):
    """
    Usuwa kopertę z bazy danych (ADMIN).
    """
    result = db.delete_envelope(envelope_id)
    
    if result["success"]:
        return jsonify(result)
    else:
        status_code = result.get("status", 400)
        return jsonify(result), status_code

@app.route('/api/envelopes/<path:envelope_id>/release', methods=['POST'])
def release_envelope(envelope_id):
    """
    Zwalnia kopertę z maszyny.
    Logika:
    - PALLETIZING -> CART-RET-05 (gotowe do zwrotu do magazynu)
    - Inne maszyny -> SHOP_FLOOR (pozostaje na hali do dalszej obróbki)
    """
    data = request.get_json() or {}
    
    result = db.release_envelope(envelope_id)
    
    if result.get("success"):
        return jsonify({
            "success": True,
            "message": f"Koperta {envelope_id} zwolniona ({result['new_status']})",
            "status": result['new_status'],
            "location": result.get('new_holder'),
            "machine": result.get('new_holder')
        })
    else:
        status_code = result.pop('status', 400)
        return jsonify(result), status_code

@app.route('/api/envelopes/<path:envelope_id>/return', methods=['POST'])
def return_envelope_to_warehouse(envelope_id):
    """
    Przyjmuje kopertę na magazyn (np. ze zwrotu lub nowej dostawy).
    Zmiana statusu: CART-RET-05 (lub inny) -> MAGAZYN.
    Body: { "location": "Sekcja A" }
    """
    data = request.get_json() or {}
    location = data.get('location', 'Sekcja A')
    
    result = db.return_to_warehouse(envelope_id, location)
    
    if result.get("success"):
        return jsonify({
            "success": True, 
            "message": f"Koperta {envelope_id} przyjęta na magazyn ({location})",
            "status": "MAGAZYN",
            "location": location
        })
    else:
        status_code = result.pop('status', 400)
        return jsonify(result), status_code

@app.route('/api/envelopes/<path:envelope_id>/notes', methods=['GET', 'POST'])
def envelope_notes(envelope_id):
    """STARY endpoint - zachowany dla kompatybilności."""
    conn = db.get_connection()
    cursor = conn.cursor()
    
    if request.method == 'GET':
        cursor.execute("""
            SELECT * FROM machine_notes 
            WHERE product_id = ?
            ORDER BY created_at DESC
            LIMIT 10
        """, (envelope_id,))
        rows = cursor.fetchall()
        conn.close()
        
        notes = []
        for row in rows:
            notes.append({
                "id": row['id'],
                "machine": row['machine_id'],
                "operator": row['operator_name'],
                "description": row['description'],
                "created_at": row['created_at']
            })
        return jsonify(notes)
    
    else:  # POST
        data = request.get_json()
        cursor.execute("""
            INSERT INTO machine_notes (machine_id, product_id, operator_name, description)
            VALUES (?, ?, ?, ?)
        """, (data.get('machine'), envelope_id, data.get('operator'), data.get('text')))
        conn.commit()
        conn.close()
        return jsonify({"success": True})

@app.route('/api/envelopes/<path:envelope_id>/history', methods=['GET'])
def get_envelope_history(envelope_id):
    """
    Pobiera historię zdarzeń dla danej koperty.
    Query params: limit (default: 5)
    
    Returns: Lista zdarzeń w formacie JSON, posortowana od najnowszych do najstarszych.
    """
    limit = request.args.get('limit', 5, type=int)
    
    history = db.get_envelope_history(envelope_id, limit)
    
    if not history:
        # Sprawdź czy koperta w ogóle istnieje
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT unique_key FROM envelopes WHERE unique_key = ?", (envelope_id,))
        exists = cursor.fetchone()
        conn.close()
        
        if not exists:
            return jsonify({"error": "Koperta nie znaleziona"}), 404
        
        # Koperta istnieje, ale nie ma historii
        return jsonify([])
    
    return jsonify(history)

# ========================
# NOWE ENDPOINTY NOTATEK PRODUKT + MASZYNA
# ========================

def get_product_code(envelope_id):
    """
    Wydobywa kod produktu z ID koperty.
    Np. '333444555#1.3#08' -> '333444555#1.3'
    """
    parts = envelope_id.split('#')
    if len(parts) >= 2:
        return f"{parts[0]}#{parts[1]}"
    return envelope_id

@app.route('/api/product-notes/<path:product_code>/<machine_id>', methods=['GET', 'POST', 'DELETE'])
def handle_product_machine_note(product_code, machine_id):
    """
    Obsługuje operacje na notatkach produkt+maszyna (GET, POST, DELETE).
    Dla POST: tworzy lub aktualizuje notatkę.
    Dla DELETE: usuwa (soft-delete).
    """
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # --- GET ---
    if request.method == 'GET':
        # Pobierz notatkę specyficzną dla tej maszyny
        cursor.execute("""
            SELECT id, product_code, machine_id, note_content, note_type,
                   created_at, modified_at, created_by, modified_by
            FROM product_machine_notes 
            WHERE product_code = ? AND machine_id = ? AND is_active = 1
            ORDER BY note_type ASC
        """, (product_code, machine_id))
        specific = cursor.fetchone()
        
        # Pobierz notatkę globalną (jeśli istnieje)
        cursor.execute("""
            SELECT id, product_code, machine_id, note_content, note_type,
                   created_at, modified_at, created_by, modified_by
            FROM product_machine_notes 
            WHERE product_code = ? AND note_type = 'global' AND is_active = 1
            LIMIT 1
        """, (product_code,))
        global_note = cursor.fetchone()
        conn.close()
        
        result = {
            "product_code": product_code,
            "machine_id": machine_id,
            "specific_note": None,
            "global_note": None
        }
        
        if specific:
            specific_images = db.get_note_images('product_machine_note', specific['id'])
            result["specific_note"] = {
                "id": specific['id'],
                "content": specific['note_content'],
                "created_at": specific['created_at'],
                "modified_at": specific['modified_at'],
                "created_by": specific['created_by'],
                "modified_by": specific['modified_by'],
                "images": [_serialize_image_meta(img) for img in specific_images]
            }
        
        if global_note:
            global_images = db.get_note_images('product_machine_note', global_note['id'])
            result["global_note"] = {
                "id": global_note['id'],
                "content": global_note['note_content'],
                "created_at": global_note['created_at'],
                "modified_at": global_note['modified_at'],
                "created_by": global_note['created_by'],
                "modified_by": global_note['modified_by'],
                "images": [_serialize_image_meta(img) for img in global_images]
            }
        
        return jsonify(result)
    
    # --- POST ---
    elif request.method == 'POST':
        data = request.get_json()
        content = data.get('content', '')
        user = data.get('user', 'Nieznany')
        note_type = data.get('note_type', 'specific')
        
        # Sprawdź czy notatka już istnieje
        cursor.execute("""
            SELECT id, note_content FROM product_machine_notes 
            WHERE product_code = ? AND machine_id = ? AND note_type = ?
        """, (product_code, machine_id, note_type))
        existing = cursor.fetchone()
        
        if existing:
            # Aktualizuj istniejącą notatkę
            old_content = existing['note_content']
            note_id = existing['id']
            
            cursor.execute("""
                UPDATE product_machine_notes 
                SET note_content = ?,
                    modified_at = CURRENT_TIMESTAMP,
                    modified_by = ?,
                    is_active = 1
                WHERE id = ?
            """, (content, user, note_id))
            
            # Zapisz historię zmiany
            cursor.execute("""
                INSERT INTO notes_history 
                (note_id, product_code, machine_id, old_content, new_content, changed_by, operation_type)
                VALUES (?, ?, ?, ?, ?, ?, 'edit')
            """, (note_id, product_code, machine_id, old_content, content, user))
            
            message = "Notatka zaktualizowana"
        else:
            # Utwórz nową notatkę
            cursor.execute("""
                INSERT INTO product_machine_notes 
                (product_code, machine_id, note_content, note_type, created_by, modified_by)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (product_code, machine_id, content, note_type, user, user))
            
            note_id = cursor.lastrowid
            
            # Zapisz historię utworzenia
            cursor.execute("""
                INSERT INTO notes_history 
                (note_id, product_code, machine_id, old_content, new_content, changed_by, operation_type)
                VALUES (?, ?, ?, NULL, ?, ?, 'create')
            """, (note_id, product_code, machine_id, content, user))
            
            message = "Notatka utworzona"
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": message,
            "note_id": note_id,
            "product_code": product_code,
            "machine_id": machine_id
        })
    
    # --- DELETE ---
    elif request.method == 'DELETE':
        data = request.get_json() or {}
        user = data.get('user', 'Nieznany')
        note_type = data.get('note_type', 'specific')
        
        cursor.execute("""
            SELECT id, note_content FROM product_machine_notes 
            WHERE product_code = ? AND machine_id = ? AND note_type = ? AND is_active = 1
        """, (product_code, machine_id, note_type))
        existing = cursor.fetchone()
        
        if not existing:
            conn.close()
            return jsonify({"error": "Notatka nie istnieje"}), 404
        
        note_id = existing['id']
        old_content = existing['note_content']
        
        # Soft delete
        cursor.execute("""
            UPDATE product_machine_notes SET is_active = 0 WHERE id = ?
        """, (note_id,))
        
        # Zapisz historię usunięcia
        cursor.execute("""
            INSERT INTO notes_history 
            (note_id, product_code, machine_id, old_content, new_content, changed_by, operation_type)
            VALUES (?, ?, ?, ?, NULL, ?, 'delete')
        """, (note_id, product_code, machine_id, old_content, user))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Notatka usunięta"
        })


@app.route('/api/operator-notes/<path:envelope_id>/<path:machine_id>', methods=['GET'])
def get_operator_notes(envelope_id, machine_id):
    limit = request.args.get('limit', 20, type=int)
    limit = max(1, min(limit, 100))
    cursor_raw = request.args.get('cursor')
    cursor = _parse_note_cursor(cursor_raw)
    if cursor_raw and cursor is None:
        return jsonify({"success": False, "error": "Nieprawidlowy cursor"}), 400

    result = db.get_operator_notes_paginated(envelope_id, machine_id, limit, cursor)
    if not result.get("success"):
        return jsonify(result), result.get("status", 500)

    notes = []
    for note in result["notes"]:
        images = db.get_note_images('operator_note', note["id"])
        notes.append(
            {
                "id": note["id"],
                "envelope_id": note["envelope_id"],
                "machine_id": note["machine_id"],
                "note_kind": note["note_kind"],
                "note_data": note["note_data"],
                "created_by": note["created_by"],
                "created_at": note["created_at"],
                "modified_at": note["modified_at"],
                "images": [_serialize_image_meta(img) for img in images],
            }
        )

    return jsonify(
        {
            "success": True,
            "notes": notes,
            "next_cursor": result.get("next_cursor"),
        }
    )


@app.route('/api/operator-notes', methods=['POST'])
def create_operator_note():
    data = request.get_json() or {}
    envelope_id = str(data.get('envelope_id', '')).strip()
    machine_id = str(data.get('machine_id', '')).strip()
    note_kind = str(data.get('note_kind', 'standard')).strip().lower()
    note_data = data.get('note_data_json') or {}
    author = str(data.get('author', 'Nieznany')).strip()

    if note_kind not in {'standard', 'pallet', 'slot'}:
        return jsonify({"success": False, "error": "Nieprawidlowy note_kind"}), 400
    if not envelope_id or not machine_id:
        return jsonify({"success": False, "error": "Wymagane pola: envelope_id, machine_id"}), 400
    if not isinstance(note_data, dict):
        return jsonify({"success": False, "error": "note_data_json musi byc obiektem"}), 400

    result = db.create_operator_note(envelope_id, machine_id, note_kind, note_data, author)
    if not result.get("success"):
        return jsonify(result), result.get("status", 500)

    return jsonify(
        {
            "success": True,
            "note_id": result["note_id"],
            "envelope_id": envelope_id,
            "machine_id": machine_id,
            "note_kind": note_kind,
        }
    )


@app.route('/api/operator-notes/<int:note_id>', methods=['DELETE'])
def delete_operator_note(note_id):
    result = db.soft_delete_operator_note(note_id)
    if not result.get("success"):
        return jsonify(result), result.get("status", 500)
    return jsonify({"success": True})


@app.errorhandler(413)
def handle_request_too_large(_error):
    _increment_note_metric("upload_fail")
    _increment_note_metric("upload_413")
    return jsonify({"success": False, "error": "Plik przekracza limit rozmiaru"}), 413


@app.route('/api/note-images', methods=['POST'])
def upload_note_image():
    uploaded_by = (request.form.get('uploaded_by') or 'anonymous').strip()
    remote_ip = request.remote_addr or 'unknown'
    if _is_rate_limited(f"user:{uploaded_by}", 20, 60):
        _increment_note_metric("upload_fail")
        _increment_note_metric("upload_429")
        return jsonify({"success": False, "error": "Rate limit user exceeded"}), 429, {"Retry-After": "60"}
    if _is_rate_limited(f"ip:{remote_ip}", 60, 60):
        _increment_note_metric("upload_fail")
        _increment_note_metric("upload_429")
        return jsonify({"success": False, "error": "Rate limit ip exceeded"}), 429, {"Retry-After": "60"}

    note_scope = (request.form.get('note_scope') or '').strip()
    if note_scope not in {'operator_note', 'product_machine_note'}:
        _increment_note_metric("upload_fail")
        return jsonify({"success": False, "error": "Nieprawidlowy note_scope"}), 400

    note_id_raw = request.form.get('note_id')
    try:
        note_id = int(note_id_raw)
    except Exception:
        _increment_note_metric("upload_fail")
        return jsonify({"success": False, "error": "Nieprawidlowy note_id"}), 400

    if not db.note_exists(note_scope, note_id):
        _increment_note_metric("upload_fail")
        return jsonify({"success": False, "error": "Notatka nie istnieje"}), 404

    order_index = request.form.get('order_index', '0')
    try:
        order_index_int = int(order_index)
    except Exception:
        order_index_int = 0

    annotations_raw = request.form.get('annotations_json')
    annotations_json = {"objects": []}
    if annotations_raw:
        try:
            annotations_json = json.loads(annotations_raw)
            if not isinstance(annotations_json, dict):
                raise ValueError("annotations_json must be object")
        except Exception:
            _increment_note_metric("upload_fail")
            return jsonify({"success": False, "error": "Nieprawidlowe annotations_json"}), 400

    file = request.files.get('file')
    if not file:
        _increment_note_metric("upload_fail")
        return jsonify({"success": False, "error": "Brak pliku"}), 400

    normalized, err, status_code = _normalize_image_to_webp(file)
    if err:
        _increment_note_metric("upload_fail")
        if status_code == 413:
            _increment_note_metric("upload_413")
        return jsonify({"success": False, "error": err}), status_code

    now = datetime.utcnow()
    rel_dir = Path(str(now.year), f"{now.month:02d}")
    abs_dir = NOTE_IMAGES_DIR / rel_dir
    abs_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid.uuid4().hex}.webp"
    rel_path = str((rel_dir / filename).as_posix())
    abs_path = abs_dir / filename

    with open(abs_path, 'wb') as f:
        f.write(normalized["bytes"])

    result = db.create_note_image(
        note_scope=note_scope,
        note_id=note_id,
        storage_path=rel_path,
        original_filename=file.filename,
        mime_type=normalized["mime_type"],
        width=normalized["width"],
        height=normalized["height"],
        size_bytes=normalized["size_bytes"],
        sha256_hex=normalized["sha256"],
        annotations_json=annotations_json,
        order_index=order_index_int,
        created_by=uploaded_by,
    )

    if not result.get("success"):
        _increment_note_metric("upload_fail")
        try:
            if abs_path.exists():
                abs_path.unlink()
        except Exception:
            pass
        return jsonify(result), result.get("status", 500)

    image_row = db.get_note_image_by_id(result["image_id"])
    _increment_note_metric("upload_success")
    return jsonify({"success": True, "image": _serialize_image_meta(image_row)})


@app.route('/api/note-images/<int:image_id>/annotations', methods=['PUT'])
def update_note_image_annotations(image_id):
    data = request.get_json() or {}
    modified_by = str(data.get('modified_by', 'anonymous')).strip()
    annotations_json = data.get('annotations_json')
    expected_revision = data.get('expected_revision')

    if not isinstance(annotations_json, dict):
        return jsonify({"success": False, "error": "annotations_json musi byc obiektem"}), 400
    try:
        expected_revision_int = int(expected_revision)
    except Exception:
        return jsonify({"success": False, "error": "expected_revision jest wymagany"}), 400

    result = db.update_note_image_annotations(image_id, annotations_json, modified_by, expected_revision_int)
    if not result.get("success"):
        if result.get("status") == 409:
            _increment_note_metric("annotations_conflict_409")
            return jsonify(result), 409
        return jsonify(result), result.get("status", 500)

    image_row = db.get_note_image_by_id(image_id)
    return jsonify({"success": True, "image": _serialize_image_meta(image_row)})


@app.route('/api/note-images/<int:image_id>', methods=['DELETE'])
def delete_note_image(image_id):
    image_row = db.get_note_image_by_id(image_id)
    if not image_row or not image_row.get("is_active"):
        return jsonify({"success": False, "error": "Obraz nie istnieje"}), 404

    result = db.soft_delete_note_image(image_id)
    if not result.get("success"):
        return jsonify(result), result.get("status", 500)

    abs_path = NOTE_IMAGES_DIR / image_row["storage_path"]
    try:
        if abs_path.exists():
            abs_path.unlink()
    except Exception:
        pass

    return jsonify({"success": True})


@app.route('/api/note-images/<int:image_id>/signed-url', methods=['GET'])
def get_note_image_signed_url(image_id):
    image_row = db.get_note_image_by_id(image_id)
    if not image_row or not image_row.get("is_active"):
        return jsonify({"success": False, "error": "Obraz nie istnieje"}), 404

    exp = int(time.time()) + SIGNED_URL_TTL_SECONDS
    return jsonify(
        {
            "success": True,
            "url": _build_signed_image_url(image_id, exp),
            "expires_at": exp,
            "etag": image_row.get("sha256"),
        }
    )


@app.route('/api/note-images/<int:image_id>/file', methods=['GET'])
def serve_note_image_file(image_id):
    exp = request.args.get('exp')
    sig = request.args.get('sig')
    if not exp or not sig or not _verify_signed_image_url(image_id, exp, sig):
        return jsonify({"success": False, "error": "Forbidden"}), 403

    image_row = db.get_note_image_by_id(image_id)
    if not image_row or not image_row.get("is_active"):
        return jsonify({"success": False, "error": "Obraz nie istnieje"}), 404

    abs_path = NOTE_IMAGES_DIR / image_row["storage_path"]
    if not abs_path.exists():
        return jsonify({"success": False, "error": "Plik nie istnieje"}), 404

    etag = image_row.get("sha256") or ""
    if_none_match = (request.headers.get("If-None-Match") or "").strip('"')
    if etag and if_none_match == etag:
        return Response(status=304, headers={"ETag": f'"{etag}"'})

    with open(abs_path, 'rb') as f:
        file_bytes = f.read()

    headers = {
        "ETag": f'"{etag}"' if etag else "",
        "Cache-Control": f"private, max-age={SIGNED_URL_TTL_SECONDS}",
        "X-Content-Type-Options": "nosniff",
        "Content-Security-Policy": "default-src 'none'",
    }
    if not headers["ETag"]:
        headers.pop("ETag")

    return Response(file_bytes, mimetype=image_row.get("mime_type") or "image/webp", headers=headers)


@app.route('/api/note-images/metrics', methods=['GET'])
def get_note_image_metrics():
    return jsonify(
        {
            "success": True,
            "metrics": {
                "upload_success": _note_image_metrics.get("upload_success", 0),
                "upload_fail": _note_image_metrics.get("upload_fail", 0),
                "upload_429": _note_image_metrics.get("upload_429", 0),
                "upload_413": _note_image_metrics.get("upload_413", 0),
                "annotations_conflict_409": _note_image_metrics.get("annotations_conflict_409", 0),
            },
        }
    )

# ========================
# ENDPOINTY PRODUKTÓW
# ========================

@app.route('/api/products', methods=['GET'])
def get_all_products():
    """Pobiera listę wszystkich produktów."""
    products = db.get_all_products()
    return jsonify(products)

@app.route('/api/products', methods=['POST'])
def create_product():
    """
    Tworzy nowy produkt.
    Body JSON:
    {
        "company_name": "PROTEGA GLOBAL LTD",
        "product_name": "T22684 OXED (NEW FSC)",
        "rcs_id": "RCS044563/C"
    }
    """
    data = request.get_json() or {}
    company_name = data.get('company_name')
    product_name = data.get('product_name')
    rcs_id = data.get('rcs_id')
    
    if not all([company_name, product_name, rcs_id]):
        return jsonify({
            "error": "Wymagane pola: company_name, product_name, rcs_id"
        }), 400
    
    result = db.create_product(company_name, product_name, rcs_id)
    
    if result.get("success"):
        return jsonify(result), 201
    else:
        return jsonify(result), 400

@app.route('/api/products/<int:product_id>', methods=['GET'])
def get_product(product_id):
    """Pobiera szczegóły produktu po ID."""
    product = db.get_product_by_id(product_id)
    if product:
        return jsonify(product)
    else:
        return jsonify({"error": "Produkt nie znaleziony"}), 404

@app.route('/api/products/search', methods=['GET'])
def search_products():
    """
    Wyszukuje produkty po fragmencie nazwy firmy, produktu lub RCS.
    Query param: q (string)
    """
    query = request.args.get('q', '')
    
    if len(query) < 2:
        return jsonify([])
    
    products = db.search_products(query)
    return jsonify(products)

@app.route('/api/products/by-rcs/<rcs_id>', methods=['GET'])
def get_product_by_rcs(rcs_id):
    """Pobiera produkt po identyfikatorze RCS."""
    product = db.get_product_by_rcs(rcs_id)
    if product:
        return jsonify(product)
    else:
        return jsonify({"error": "Produkt nie znaleziony"}), 404

@app.route('/api/products/<int:product_id>', methods=['PUT'])
def update_product(product_id):
    """
    Aktualizuje dane produktu.
    Body JSON:
    {
        "company_name": "PROTEGA GLOBAL LTD",
        "product_name": "T22684 OXED (NEW FSC)",
        "rcs_id": "RCS044563/C"
    }
    """
    data = request.get_json() or {}
    
    company_name = data.get('company_name')
    product_name = data.get('product_name')
    rcs_id = data.get('rcs_id')
    
    result = db.update_product(product_id, company_name, product_name, rcs_id)
    
    if result.get("success"):
        return jsonify(result)
    else:
        return jsonify(result), result.get('status', 400)

@app.route('/api/products/<int:product_id>', methods=['DELETE'])
def delete_product(product_id):
    """
    Usuwa produkt (soft delete - ustawia is_active = 0).
    Sprawdza czy produkt nie jest używany przez koperty.
    """
    result = db.delete_product_soft(product_id)
    
    if result.get("success"):
        return jsonify(result)
    else:
        return jsonify(result), result.get('status', 400)

@app.route('/api/products/import-csv', methods=['POST'])
def import_products_csv():
    """
    Importuje produkty z pliku CSV.
    Oczekuje pliku w formacie multipart/form-data z kluczem 'file'.
    
    Format CSV (z nagłówkami):
    company_name,product_name,rcs_id
    """
    if 'file' not in request.files:
        return jsonify({"error": "Brak pliku"}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"error": "Nie wybrano pliku"}), 400
    
    if not file.filename.endswith('.csv'):
        return jsonify({"error": "Plik musi być w formacie CSV"}), 400
    
    try:
        import csv
        import io
        
        # Odczytaj zawartość pliku
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        products_to_import = []
        for row in csv_reader:
            if 'company_name' in row and 'product_name' in row and 'rcs_id' in row:
                products_to_import.append({
                    'company_name': row['company_name'].strip(),
                    'product_name': row['product_name'].strip(),
                    'rcs_id': row['rcs_id'].strip()
                })
        
        if not products_to_import:
            return jsonify({"error": "Plik CSV jest pusty lub ma nieprawidłowy format"}), 400
        
        # Importuj produkty
        result = db.import_products_batch(products_to_import)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": f"Błąd przetwarzania pliku CSV: {str(e)}"}), 500

@app.route('/api/products/import-excel', methods=['POST'])
def import_products_excel():
    """
    Importuje produkty z pliku Excel (.xlsx, .xls).
    Oczekuje pliku w formacie multipart/form-data z kluczem 'file'.
    
    Format Excel (pierwsza kolumna to nagłówki):
    company_name | product_name | rcs_id
    """
    if 'file' not in request.files:
        return jsonify({"error": "Brak pliku"}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"error": "Nie wybrano pliku"}), 400
    
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({"error": "Plik musi być w formacie Excel (.xlsx lub .xls)"}), 400
    
    try:
        import openpyxl
        import io
        
        # Odczytaj plik Excel
        workbook = openpyxl.load_workbook(io.BytesIO(file.read()))
        sheet = workbook.active
        
        products_to_import = []
        
        # Znajdź indeksy kolumn (pierwszy wiersz to nagłówki)
        headers = [cell.value for cell in sheet[1]]
        
        try:
            company_idx = headers.index('company_name')
            product_idx = headers.index('product_name')
            rcs_idx = headers.index('rcs_id')
        except ValueError:
            return jsonify({
                "error": "Brak wymaganych kolumn w pliku Excel. Wymagane: company_name, product_name, rcs_id"
            }), 400
        
        # Przetwórz wiersze (pomijamy pierwszy wiersz z nagłówkami)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[company_idx] and row[product_idx] and row[rcs_idx]:
                products_to_import.append({
                    'company_name': str(row[company_idx]).strip(),
                    'product_name': str(row[product_idx]).strip(),
                    'rcs_id': str(row[rcs_idx]).strip()
                })
        
        if not products_to_import:
            return jsonify({"error": "Plik Excel jest pusty lub nie zawiera danych"}), 400
        
        # Importuj produkty
        result = db.import_products_batch(products_to_import)
        return jsonify(result)
        
    except ImportError:
        return jsonify({
            "error": "Biblioteka openpyxl nie jest zainstalowana. Zainstaluj: pip install openpyxl"
        }), 500
    except Exception as e:
        return jsonify({"error": f"Błąd przetwarzania pliku Excel: {str(e)}"}), 500

# ========================
# WAREHOUSE SEARCH LIST API
# ========================

@app.route('/api/search-list/today', methods=['GET'])
def get_todays_search_list():
    """Pobiera dzisiejszą listę wyszukiwania."""
    user_id = request.args.get('user_id')  # Opcjonalne - dla list per-user
    items = db.get_todays_search_list(user_id)
    return jsonify(items)

@app.route('/api/search-list/add', methods=['POST'])
def add_to_search_list():
    """Dodaje pojedynczą kopertę do listy wyszukiwania."""
    data = request.get_json()
    envelope_id = data.get('envelope_id')
    user_id = data.get('user_id')
    priority = data.get('priority', 'normal')
    
    if not envelope_id:
        return jsonify({"error": "envelope_id jest wymagane"}), 400
    
    result = db.add_to_search_list(envelope_id, user_id, priority)
    status = result.get('status', 200 if result['success'] else 400)
    return jsonify(result), status

@app.route('/api/search-list/bulk', methods=['POST'])
def bulk_add_search_list():
    """Dodaje wiele kopert naraz (text/plain - jedna koperta na linię)."""
    user_id = request.args.get('user_id')
    
    # Obsługa tekstu (jedna koperta na linię)
    if request.content_type == 'text/plain':
        text = request.get_data(as_text=True)
        envelope_ids = [line.strip() for line in text.split('\n') if line.strip()]
        
        result = db.bulk_add_to_search_list(envelope_ids, user_id)
        return jsonify(result)
    
    # Obsługa JSON
    data = request.get_json()
    envelope_ids = data.get('envelope_ids', [])
    
    if not envelope_ids:
        return jsonify({"error": "envelope_ids (array) jest wymagane"}), 400
    
    result = db.bulk_add_to_search_list(envelope_ids, user_id)
    return jsonify(result)

@app.route('/api/search-list/import-csv', methods=['POST'])
def import_search_list_csv():
    """Import listy wyszukiwania z CSV (1 kolumna - envelope ID)."""
    if 'file' not in request.files:
        return jsonify({"error": "Brak pliku"}), 400
    
    file = request.files['file']
    user_id = request.args.get('user_id')
    
    try:
        import csv
        from io import StringIO
        
        content = file.read().decode('utf-8')
        envelope_ids = []
        
        # Próbuj jako CSV
        csv_reader = csv.reader(StringIO(content))
        for row in csv_reader:
            if row and row[0].strip():
                envelope_ids.append(row[0].strip())
        
        result = db.bulk_add_to_search_list(envelope_ids, user_id)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": f"Błąd przetwarzania CSV: {str(e)}"}), 500

@app.route('/api/search-list/import-excel', methods=['POST'])
def import_search_list_excel():
    """Import listy wyszukiwania z Excel (1 kolumna - envelope ID)."""
    if 'file' not in request.files:
        return jsonify({"error": "Brak pliku"}), 400
    
    file = request.files['file']
    user_id = request.args.get('user_id')
    
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active
        
        envelope_ids = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                envelope_ids.append(str(row[0]).strip())
        
        result = db.bulk_add_to_search_list(envelope_ids, user_id)
        return jsonify(result)
        
    except ImportError:
        return jsonify({"error": "Biblioteka openpyxl nie jest zainstalowana"}), 500
    except Exception as e:
        return jsonify({"error": f"Błąd przetwarzania Excel: {str(e)}"}), 500

@app.route('/api/search-list/mark-found/<envelope_id>', methods=['PUT'])
def mark_search_item_found(envelope_id):
    """Oznacza kopertę jako znalezioną."""
    result = db.mark_search_item_found(envelope_id)
    return jsonify(result)

@app.route('/api/search-list/<int:item_id>', methods=['DELETE'])
def delete_search_item(item_id):
    """Usuwa element z listy wyszukiwania."""
    result = db.delete_search_item(item_id)
    return jsonify(result)

@app.route('/api/search-list/clear', methods=['DELETE'])
def clear_search_list():
    """Czyści dzisiejszą listę wyszukiwania."""
    user_id = request.args.get('user_id')
    result = db.clear_todays_search_list(user_id)
    return jsonify(result)

# ========================
# URUCHOMIENIE
# ========================

@app.route('/api/envelopes/<path:envelope_id>/history', methods=['GET'])
def get_envelope_history_api(envelope_id):
    """Pobiera historię zdarzeń dla koperty."""
    limit = request.args.get('limit', 5, type=int)
    history = db.get_envelope_history(envelope_id, limit)
    return jsonify(history)

@app.route('/api/envelopes', methods=['POST'])
def create_envelope_api():
    """
    Tworzy nową kopertę w systemie.
    Body JSON:
    {
        "product_id": 1,              -- ID produktu z tabeli products
        "warehouse_section": "A"      -- Sekcja magazynowa (opcjonalne)
    }
    
    Alternatywnie (stary format dla kompatybilności):
    {
        "rcs_id": "123456789",
        "warehouse_section": "A"
    }
    """
    data = request.get_json() or {}
    product_id = data.get('product_id')
    rcs_id = data.get('rcs_id')
    section = data.get('warehouse_section', 'A')
    
    # Jeśli podano product_id, pobierz rcs_id z tabeli products
    if product_id:
        product = db.get_product_by_id(product_id)
        if not product:
            return jsonify({"error": f"Produkt o ID {product_id} nie istnieje"}), 404
        rcs_id = product['rcs_id']
    
    if not rcs_id:
        return jsonify({"error": "Wymagane pola: product_id (lub rcs_id)"}), 400
        
    # Format ID: rcs_id (bez numeru kopii)
    unique_key = rcs_id
    
    conn = db.get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO envelopes 
            (unique_key, rcs_id, status, 
             current_holder_id, current_holder_type, warehouse_section, creation_reason, product_id)
            VALUES (?, ?, 'MAGAZYN', ?, 'WAREHOUSE', ?, 'MANUAL', ?)
        """, (unique_key, rcs_id, f"MAGAZYN_{section}", section, product_id))

        # Dodaj wpis do historii
        cursor.execute("""
            INSERT INTO events 
            (envelope_key, user_id, from_status, to_status, from_holder, to_holder, comment)
            VALUES (?, 'ADMIN', 'NEW', 'MAGAZYN', 'SYSTEM', 'WAREHOUSE', 'Utworzono ręcznie w panelu admina')
        """, (unique_key,))
        
        conn.commit()
        success = True
    except Exception as e:
        success = False
        error_msg = str(e)
    finally:
        conn.close()
        
    if success:
        return jsonify({
            "success": True,
            "message": f"Utworzono kopertę {unique_key}",
            "id": unique_key
        }), 201
    else:
        return jsonify({"error": f"Błąd tworzenia koperty (może duplikat?): {error_msg}"}), 400

@app.route('/api/stats/cart-return-count', methods=['GET'])
def get_cart_return_count():
    conn = db.get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM envelopes WHERE status = 'CART-RET-05'")
        result = cursor.fetchone()
        count = result[0] if result else 0
    except Exception as e:
        print(f"Error counting cart return: {e}")
        count = 0
    finally:
        conn.close()
    return jsonify({"count": count})

@app.route('/api/stats/cart-return-list', methods=['GET'])
def get_cart_return_list():
    """Zwraca listę wszystkich kopert na CART-RET-05."""
    conn = db.get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT unique_key, rcs_id, updated_at
            FROM envelopes 
            WHERE status = 'CART-RET-05'
            ORDER BY updated_at DESC
        """)
        rows = cursor.fetchall()
        
        envelopes = []
        for row in rows:
            envelopes.append({
                "unique_key": row['unique_key'],
                "rcs_id": row['rcs_id'] or '',
                "updated_at": row['updated_at']
            })
        
        return jsonify(envelopes)
    except Exception as e:
        print(f"Error fetching cart return list: {e}")
        return jsonify([])
    finally:
        conn.close()

@app.route('/api/machines/<path:machine_id>/status', methods=['GET'])
def get_machine_status(machine_id):
    """Sprawdza status maszyny (czy ma przypisaną kopertę)."""
    envelope_data = db.get_machine_status(machine_id)
    
    if envelope_data:
        return jsonify({
            "status": "W_PRODUKCJI",
            "data": envelope_data
        })
    else:
        return jsonify({
            "status": "IDLE",
            "data": None
        })

@app.route('/api/machines', methods=['GET'])
def get_machines():
    """Lista maszyn operatora (aktywne lub wszystkie dla admina)."""
    include_all = request.args.get('all', '0') == '1'
    machines = db.get_all_machines_auth() if include_all else db.get_active_machines_auth()
    return jsonify(machines)

@app.route('/api/machines', methods=['POST'])
def create_machine():
    """Dodaje nową maszynę z PIN."""
    data = request.get_json() or {}
    machine_name = data.get('machine')
    pin = data.get('pin')
    result = db.create_machine_auth(machine_name, pin)
    if result.get("success"):
        return jsonify(result), 201
    return jsonify(result), result.get("status", 500)

@app.route('/api/machines/<int:machine_id>', methods=['PUT'])
def update_machine(machine_id):
    """Aktualizuje nazwę i/lub aktywność maszyny."""
    data = request.get_json() or {}
    result = db.update_machine_auth(
        machine_id,
        machine_name=data.get('machine'),
        is_active=data.get('is_active')
    )
    if result.get("success"):
        return jsonify(result)
    return jsonify(result), result.get("status", 500)

@app.route('/api/machines/<int:machine_id>/pin', methods=['PUT'])
def update_machine_pin(machine_id):
    """Zmienia PIN maszyny."""
    data = request.get_json() or {}
    new_pin = data.get('new_pin')
    result = db.change_machine_pin(machine_id, new_pin)
    if result.get("success"):
        return jsonify(result)
    return jsonify(result), result.get("status", 500)


# ========================
# ZARZĄDZANIE UŻYTKOWNIKAMI
# ========================

@app.route('/api/users', methods=['GET'])
def get_users():
    """Pobiera listę wszystkich użytkowników."""
    users = db.get_all_users()
    return jsonify(users)

@app.route('/api/users', methods=['POST'])
def create_user():
    """Tworzy nowego użytkownika."""
    data = request.get_json()
    
    if not data:
        return jsonify({"success": False, "error": "Brak danych"}), 400
    
    username = data.get('username')
    pin = data.get('pin')
    role = data.get('role')
    full_name = data.get('full_name')
    shift = data.get('shift')
    
    if not username or not pin or not role:
        return jsonify({"success": False, "error": "Wymagane pola: username, pin, role"}), 400
    
    result = db.create_user(username, pin, role, full_name, shift)
    
    if result['success']:
        return jsonify(result), 201
    else:
        return jsonify(result), result.get('status', 500)

@app.route('/api/users/<int:user_id>', methods=['GET'])
def get_user(user_id):
    """Pobiera dane konkretnego użytkownika."""
    user = db.get_user_by_id(user_id)
    
    if user:
        return jsonify(user)
    else:
        return jsonify({"success": False, "error": "Użytkownik nie znaleziony"}), 404

@app.route('/api/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    """Aktualizuje dane użytkownika."""
    data = request.get_json()
    
    if not data:
        return jsonify({"success": False, "error": "Brak danych"}), 400
    
    full_name = data.get('full_name')
    role = data.get('role')
    shift = data.get('shift')
    is_active = data.get('is_active')
    
    result = db.update_user(user_id, full_name, role, shift, is_active)
    
    if result['success']:
        return jsonify(result)
    else:
        return jsonify(result), result.get('status', 500)

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    """Usuwa użytkownika (soft delete)."""
    result = db.delete_user(user_id)
    
    if result['success']:
        return jsonify(result)
    else:
        return jsonify(result), result.get('status', 500)

@app.route('/api/users/<int:user_id>/change-pin', methods=['POST'])
def change_pin(user_id):
    """Zmienia PIN użytkownika."""
    data = request.get_json()
    
    if not data or 'new_pin' not in data:
        return jsonify({"success": False, "error": "Wymagane pole: new_pin"}), 400
    
    new_pin = data['new_pin']
    result = db.change_user_pin(user_id, new_pin)
    
    if result['success']:
        return jsonify(result)
    else:
        return jsonify(result), result.get('status', 500)

@app.route('/api/auth/verify', methods=['POST'])
def verify_user():
    """Weryfikuje użytkownika po PIN."""
    data = request.get_json()
    
    if not data or 'pin' not in data:
        return jsonify({"success": False, "error": "Wymagane pole: pin"}), 400
    
    pin = data['pin']
    role = data.get('role')  # Opcjonalne filtrowanie po roli
    
    user = db.verify_user(pin, role)
    
    if user:
        return jsonify({"success": True, "user": user})
    else:
        return jsonify({"success": False, "error": "Nieprawidłowy PIN lub użytkownik nieaktywny"}), 401

@app.route('/api/auth/machine-verify', methods=['POST'])
def verify_machine():
    """Weryfikuje PIN operatora dla konkretnej maszyny."""
    data = request.get_json() or {}
    machine_name = data.get('machine')
    pin = data.get('pin')

    if not machine_name or not pin:
        return jsonify({"success": False, "error": "Wymagane pola: machine, pin"}), 400

    result = db.verify_machine_auth(machine_name, pin)
    if result.get("success"):
        return jsonify({"success": True, "machine": result["machine"]})
    return jsonify({"success": False, "error": result.get("error")}), result.get("status", 401)

# ========================
# ENDPOINTY HISTORII OBIEGU
# ========================

@app.route('/api/circulation-history/<rcs_id>', methods=['GET'])
def get_circulation_history_txt(rcs_id):
    """
    Generuje i zwraca plik TXT z historią obiegu koperty po RCS.
    """
    try:
        from circulation_history import get_envelope_circulation_history, format_circulation_history_txt
        
        history_data = get_envelope_circulation_history(rcs_id)
        
        if not history_data:
            return jsonify({"error": f"Brak danych dla RCS: {rcs_id}"}), 404
        
        txt_content = format_circulation_history_txt(rcs_id, history_data)
        
        # Zwróć jako plik do pobrania
        from flask import Response
        return Response(
            txt_content,
            mimetype='text/plain',
            headers={'Content-Disposition': f'attachment; filename=circulation_history_{rcs_id}.txt'}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/circulation-history/<rcs_id>/view', methods=['GET'])
def view_circulation_history_json(rcs_id):
    """
    Zwraca historię obiegu koperty w formacie JSON.
    """
    try:
        from circulation_history import get_envelope_circulation_history
        
        history_data = get_envelope_circulation_history(rcs_id)
        
        if not history_data:
            return jsonify({"error": f"Brak danych dla RCS: {rcs_id}"}), 404
        
        return jsonify({
            "success": True,
            "rcs_id": rcs_id,
            "history": history_data
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/circulation-history/generate-all', methods=['POST'])
def generate_all_histories():
    """
    Generuje pliki historii dla wszystkich RCS w systemie.
    """
    try:
        from circulation_history import save_all_circulation_histories
        
        saved_files = save_all_circulation_histories()
        
        return jsonify({
            "success": True,
            "message": f"Wygenerowano {len(saved_files)} plików historii",
            "files": saved_files
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    init_demo_envelopes()
    port = int(os.environ.get('PORT', 5000))
    is_production = APP_ENV == 'production'
    print(f"\n🚀 Serwer API uruchomiony na http://localhost:{port}")
    print("📄 Otwórz prototype.html w przeglądarce")
    print("   (upewnij się, że serwer działa w tle)\n")
    app.run(host='0.0.0.0', debug=not is_production, use_reloader=False, port=port)
